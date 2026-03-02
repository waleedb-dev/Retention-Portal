#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required for this script. Install it with: pip install openpyxl"
    ) from exc

API_BASE_URL = "https://services.leadconnectorhq.com"
API_VERSION = "2021-07-28"
DEFAULT_GHL_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/137.0.0.0 Safari/537.36"
)

SUBAGENT_ALIASES = {
    "ambition bpo": "ambition",
    "arktech bpo": "ark tech",
    "argon comm bpo": "argon comm",
    "cerberus bpo": "cerberus bpo",
    "corebiz bpo": "corebiz",
    "crown connect bpo": "crown connect bpo",
    "crossnotch bpo": "crossnotch",
    "digicon bpo": "digicon",
    "downtown": "downtown bpo",
    "downtown bpo": "downtown bpo",
    "everest bpo": "everest bpo",
    "everline solution bpo": "everline solution",
    "growthonics bpo": "growthonics bpo",
    "maverick communications": "maverick",
    "optimum bpo": "optimum bpo",
    "plexi bpo": "plexi",
    "pro soliutions bpo": "pro solutions bpo",
    "seller": "sellerz bpo",
    "sellerz": "sellerz bpo",
    "sellerz bpo": "sellerz bpo",
    "stratix": "stratix bpo",
    "stratix bpo": "stratix bpo",
    "the zupax marketing": "zupax marketing",
    "trust link": "trust link",
    "vize": "vize bpo",
    "vyn bpo": "vyn",
}


@dataclass
class DealRecord:
    id: int | None
    monday_item_id: str
    ghl_name: str
    call_center: str
    updated_at: str | None


@dataclass
class SubagentRecord:
    id: str
    name: str
    location_id: str
    token: str
    token_source: str


@dataclass
class ContactRecord:
    contact_name: str
    contact_id: str
    account_id: str
    updated_on: str
    row_number: int


def load_env_file(path: Path) -> None:
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key or key in os.environ:
            continue

        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]

        os.environ[key] = value


def normalize_whitespace(value: str) -> str:
    return " ".join(value.split())


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "")
    value = value.replace("\u00A0", " ")
    value = value.strip().lower()
    value = re.sub(r"[^\w\s]", " ", value)
    return normalize_whitespace(value)


def call_center_candidates(value: str) -> list[str]:
    normalized = normalize_text(value)
    if not normalized:
        return []

    candidates = [normalized]
    alias = SUBAGENT_ALIASES.get(normalized)
    if alias and alias not in candidates:
        candidates.append(alias)

    simplified = re.sub(
        r"\b(bpo|communications|communication|marketing|tech|solutions|solution|comm)\b",
        "",
        normalized,
    )
    simplified = normalize_whitespace(simplified)
    if simplified and simplified not in candidates:
        candidates.append(simplified)

    return candidates


def contact_name_candidates(value: str) -> list[str]:
    candidates: list[str] = []
    raw = normalize_text(value)
    if raw:
        candidates.append(raw)

    if "," in value:
        parts = [normalize_whitespace(part) for part in value.split(",") if part.strip()]
        if len(parts) >= 2:
            reordered = normalize_text(" ".join(parts[1:] + [parts[0]]))
            if reordered and reordered not in candidates:
                candidates.append(reordered)

    return candidates


def parse_datetime(value: str) -> tuple[int, str]:
    text = (value or "").strip()
    if not text:
        return (0, "")

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
        try:
            parsed = datetime.strptime(text, fmt)
            return (int(parsed.timestamp()), text)
        except ValueError:
            continue

    return (0, text)


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def http_json(url: str, headers: dict[str, str]) -> Any:
    request = Request(url, headers=headers)
    with urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def build_ghl_headers(token: str, version: str, location_id: str) -> dict[str, str]:
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}",
        "Version": version,
        "User-Agent": os.environ.get("GHL_API_USER_AGENT", DEFAULT_GHL_USER_AGENT),
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Origin": "https://app.gohighlevel.com",
        "Referer": "https://app.gohighlevel.com/",
    }
    if location_id:
        headers["Location-Id"] = location_id
    return headers


def fetch_one_deal(
    supabase_url: str,
    service_role_key: str,
    monday_item_id: str | None,
    offset: int,
    active_only: bool,
) -> DealRecord:
    select = "id,monday_item_id,ghl_name,call_center,updated_at"
    base = f"{supabase_url.rstrip('/')}/rest/v1/monday_com_deals?select={quote(select)}"
    if monday_item_id:
        base += f"&monday_item_id=eq.{quote(monday_item_id)}"
    else:
        if active_only:
            base += "&is_active=eq.true"
        base += "&ghl_name=not.is.null&call_center=not.is.null&order=updated_at.desc.nullslast&limit=1"
        if offset > 0:
            base += f"&offset={offset}"

    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Accept": "application/json",
    }

    print("[ghl-one] fetching one deal from Supabase")
    rows = http_json(base, headers)
    if not isinstance(rows, list) or not rows:
        raise RuntimeError("No matching monday_com_deals row found")

    row = rows[0]
    ghl_name = str(row.get("ghl_name") or "").strip()
    call_center = str(row.get("call_center") or "").strip()
    monday_value = str(row.get("monday_item_id") or "").strip()
    if not ghl_name or not call_center or not monday_value:
        raise RuntimeError("Selected deal is missing ghl_name, call_center, or monday_item_id")

    return DealRecord(
        id=row.get("id") if isinstance(row.get("id"), int) else None,
        monday_item_id=monday_value,
        ghl_name=ghl_name,
        call_center=call_center,
        updated_at=str(row.get("updated_at") or "").strip() or None,
    )


def load_subagents(path: Path) -> tuple[dict[str, SubagentRecord], list[SubagentRecord]]:
    parsed = load_json(path)
    if not isinstance(parsed, list):
        raise RuntimeError("subagent json must contain a top-level array")

    by_name: dict[str, SubagentRecord] = {}
    ordered: list[SubagentRecord] = []
    for entry in parsed:
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name") or "").strip()
        account_id = str(entry.get("id") or "").strip()
        location_id = str(entry.get("location_id") or "").strip()
        access_token = str(entry.get("access_token") or "").strip()
        api_key = str(entry.get("api_key") or "").strip()
        token = access_token or api_key
        token_source = "access_token" if access_token else "api_key"
        if not name or not account_id or not token:
            continue

        record = SubagentRecord(
            id=account_id,
            name=name,
            location_id=location_id,
            token=token,
            token_source=token_source,
        )
        ordered.append(record)
        by_name[normalize_text(name)] = record

    return by_name, ordered


def resolve_subagent(
    call_center: str,
    subagents_by_name: dict[str, SubagentRecord],
    ordered_subagents: list[SubagentRecord],
) -> tuple[SubagentRecord | None, str]:
    for candidate in call_center_candidates(call_center):
        if candidate in subagents_by_name:
            return subagents_by_name[candidate], "exact_or_alias"

    normalized_call_center = normalize_text(call_center)
    best: SubagentRecord | None = None
    best_score = -1
    for record in ordered_subagents:
        normalized_name = normalize_text(record.name)
        if not normalized_name:
            continue
        if normalized_name in normalized_call_center or normalized_call_center in normalized_name:
            score = min(len(normalized_name), len(normalized_call_center))
            if score > best_score:
                best = record
                best_score = score

    if best is not None:
        return best, "contains"

    return None, "not_found"


def choose_best_contact(records: list[ContactRecord]) -> ContactRecord:
    def sort_key(record: ContactRecord) -> tuple[int, str, int]:
        timestamp, updated_text = parse_datetime(record.updated_on)
        return (timestamp, updated_text, record.row_number)

    return sorted(records, key=sort_key, reverse=True)[0]


def load_contacts(path: Path) -> dict[tuple[str, str], ContactRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(min_row=1, values_only=True)

    headers = [str(cell or "").strip() for cell in next(rows)]
    header_index = {header: index for index, header in enumerate(headers)}
    required = {"Contact Name", "Contact ID", "Account Id", "Updated on"}
    missing = required - set(header_index)
    if missing:
        raise RuntimeError(f"workbook missing required columns: {', '.join(sorted(missing))}")

    grouped: dict[tuple[str, str], list[ContactRecord]] = {}
    row_number = 1
    for values in rows:
        row_number += 1
        contact_name = str(values[header_index["Contact Name"]] or "").strip()
        contact_id = str(values[header_index["Contact ID"]] or "").strip()
        account_id = str(values[header_index["Account Id"]] or "").strip()
        updated_on = str(values[header_index["Updated on"]] or "").strip()
        if not contact_name or not contact_id or not account_id:
            continue

        record = ContactRecord(
            contact_name=contact_name,
            contact_id=contact_id,
            account_id=account_id,
            updated_on=updated_on,
            row_number=row_number,
        )

        for candidate in contact_name_candidates(contact_name):
            grouped.setdefault((account_id, candidate), []).append(record)

    return {key: choose_best_contact(records) for key, records in grouped.items()}


def fetch_contact_notes_list(contact_id: str, token: str, location_id: str, version: str) -> Any:
    url = f"{API_BASE_URL}/contacts/{quote(contact_id)}/notes"
    return http_json(url, build_ghl_headers(token, version, location_id))


def extract_note_list(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [entry for entry in payload if isinstance(entry, dict)]
    if isinstance(payload, dict):
        for key in ("notes", "data", "contactNotes"):
            value = payload.get(key)
            notes = extract_note_list(value)
            if notes:
                return notes
    return []


def choose_note(notes: list[dict[str, Any]]) -> dict[str, Any] | None:
    def sort_key(entry: dict[str, Any]) -> tuple[int, str]:
        for key in ("updatedAt", "updated_at", "createdAt", "created_at", "dateAdded"):
            value = str(entry.get(key) or "").strip()
            timestamp, text = parse_datetime(value.replace("T", " ").replace("Z", ""))
            if timestamp or text:
                return (timestamp, text)
        return (0, "")

    if not notes:
        return None
    return sorted(notes, key=sort_key, reverse=True)[0]


def fetch_note_detail(
    contact_id: str,
    note_id: str,
    token: str,
    location_id: str,
    version: str,
) -> Any:
    url = f"{API_BASE_URL}/contacts/{quote(contact_id)}/notes/{quote(note_id)}"
    return http_json(url, build_ghl_headers(token, version, location_id))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Fetch one lead from monday_com_deals and resolve all GHL notes through subagent/workbook mapping"
    )
    parser.add_argument("--monday-item-id", help="Optional exact monday_item_id to fetch")
    parser.add_argument("--offset", type=int, default=0, help="Pick the Nth latest deal when monday_item_id is not provided")
    parser.add_argument("--active-only", action="store_true")
    parser.add_argument("--subagents-json", default="notes/subagent.json")
    parser.add_argument("--contacts-xlsx", default="notes/GHL_Opportunities_New_API_Export (19).xlsx")
    parser.add_argument("--output", default="notes/one_lead_ghl_note.json")
    parser.add_argument("--note-id", help="Optional exact note id. If provided, script also fetches that note detail after listing notes")
    parser.add_argument("--version", default=API_VERSION)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    load_env_file((Path.cwd() / ".env").resolve())

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_role_key:
        raise RuntimeError("NEXT_PUBLIC_SUPABASE_URL/SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required")

    root = Path.cwd()
    deal = fetch_one_deal(
        supabase_url=supabase_url,
        service_role_key=service_role_key,
        monday_item_id=args.monday_item_id,
        offset=args.offset,
        active_only=args.active_only,
    )
    print("[ghl-one] selected deal", {
        "id": deal.id,
        "monday_item_id": deal.monday_item_id,
        "ghl_name": deal.ghl_name,
        "call_center": deal.call_center,
    })

    subagents_by_name, ordered_subagents = load_subagents((root / args.subagents_json).resolve())
    subagent, match_mode = resolve_subagent(deal.call_center, subagents_by_name, ordered_subagents)
    if subagent is None:
        raise RuntimeError(f'No subagent match found for call center "{deal.call_center}"')
    print("[ghl-one] matched subagent", {
        "name": subagent.name,
        "account_id": subagent.id,
        "location_id": subagent.location_id,
        "match_mode": match_mode,
        "token_source": subagent.token_source,
    })

    contacts_lookup = load_contacts((root / args.contacts_xlsx).resolve())
    matched_contact: ContactRecord | None = None
    for candidate in contact_name_candidates(deal.ghl_name):
        matched_contact = contacts_lookup.get((subagent.id, candidate))
        if matched_contact is not None:
            break

    if matched_contact is None:
        raise RuntimeError(
            f'No workbook contact found for ghl_name "{deal.ghl_name}" and account_id "{subagent.id}"'
        )
    print("[ghl-one] matched contact", {
        "contact_name": matched_contact.contact_name,
        "contact_id": matched_contact.contact_id,
        "account_id": matched_contact.account_id,
    })

    notes_list_payload = fetch_contact_notes_list(
        contact_id=matched_contact.contact_id,
        token=subagent.token,
        location_id=subagent.location_id,
        version=args.version,
    )
    notes = extract_note_list(notes_list_payload)
    print(f"[ghl-one] fetched notes list count={len(notes)}")

    latest_note = choose_note(notes)
    latest_note_id = str(latest_note.get("id") or "") if latest_note else ""
    note_detail = None
    requested_note_id = (args.note_id or "").strip()
    requested_note_summary = None

    if requested_note_id:
        requested_note_summary = next(
            (note for note in notes if str(note.get("id") or "") == requested_note_id),
            None,
        )
        print("[ghl-one] fetching note detail", {
            "note_id": requested_note_id,
        })
        note_detail = fetch_note_detail(
            contact_id=matched_contact.contact_id,
            note_id=requested_note_id,
            token=subagent.token,
            location_id=subagent.location_id,
            version=args.version,
        )

    output = {
        "deal": {
            "id": deal.id,
            "monday_item_id": deal.monday_item_id,
            "ghl_name": deal.ghl_name,
            "call_center": deal.call_center,
            "updated_at": deal.updated_at,
        },
        "subagent": {
            "id": subagent.id,
            "name": subagent.name,
            "location_id": subagent.location_id,
            "match_mode": match_mode,
            "token_source": subagent.token_source,
        },
        "contact": {
            "account_id": matched_contact.account_id,
            "contact_name": matched_contact.contact_name,
            "contact_id": matched_contact.contact_id,
            "updated_on": matched_contact.updated_on,
        },
        "notes_list_count": len(notes),
        "latest_note_id": latest_note_id or None,
        "latest_note_summary": latest_note,
        "requested_note_id": requested_note_id or None,
        "requested_note_summary": requested_note_summary,
        "notes_list_payload": notes_list_payload,
        "notes": notes,
        "note_detail": note_detail,
    }

    output_path = (root / args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"[ghl-one] wrote {output_path}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        print(f"[ghl-one] HTTP {exc.code}: {body}")
        sys.exit(1)
    except URLError as exc:
        print(f"[ghl-one] URL error: {exc}")
        sys.exit(1)
    except Exception as exc:
        print(f"[ghl-one] failed: {exc}")
        sys.exit(1)
