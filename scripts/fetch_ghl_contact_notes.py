#!/usr/bin/env python3
"""
Fetch GHL contact notes for deals by joining:
1) monday_com_deals (Supabase) -> ghl_name + call_center
2) notes/subagent.json -> subagent account id + token by call_center
3) notes/GHL_Opportunities_New_API_Export (19).xlsx -> contact_id by Contact Name + Account Id
4) GHL contact notes API -> notes by contact_id + token

Defaults:
- subagents JSON: notes/subagent.json
- opportunities XLSX: notes/GHL_Opportunities_New_API_Export (19).xlsx
- output JSON: notes/deal_contact_notes.json
- output CSV: notes/deal_contact_notes.csv

Notes:
- Uses access_token first, then api_key as fallback
- Supports --deals-csv for offline/dry-run testing without Supabase access
- Supports --dry-run to skip the external notes API and only validate joins
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - runtime dependency check
    raise SystemExit(
        "openpyxl is required for this script. Install it with: pip install openpyxl"
    ) from exc


SUBAGENT_ALIASES = {
    "ambition bpo": "ambition",
    "arktech bpo": "ark tech",
    "argon comm bpo": "argon comm",
    "cerberus bpo": "cerberus bpo",
    "corebiz bpo": "corebiz",
    "crown connect bpo": "crown connect bpo",
    "crossnotch bpo": "crossnotch",
    "digicon bpo": "digicon",
    "downtown": "downtown bpo",
    "downtown bpo": "downtown bpo",
    "everest bpo": "everest bpo",
    "everline solution bpo": "everline solution",
    "growthonics bpo": "growthonics bpo",
    "maverick communications": "maverick",
    "optimum bpo": "optimum bpo",
    "plexi bpo": "plexi",
    "pro soliutions bpo": "pro solutions bpo",
    "seller": "sellerz bpo",
    "sellerz": "sellerz bpo",
    "sellerz bpo": "sellerz bpo",
    "stratix": "stratix bpo",
    "stratix bpo": "stratix bpo",
    "the zupax marketing": "zupax marketing",
    "trust link": "trust link",
    "vize": "vize bpo",
    "vyn bpo": "vyn",
}
DEFAULT_GHL_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/137.0.0.0 Safari/537.36"
)


@dataclass
class DealRecord:
    monday_item_id: str
    ghl_name: str
    call_center: str


@dataclass
class SubagentRecord:
    id: str
    name: str
    location_id: str
    token: str
    token_source: str


@dataclass
class ContactRecord:
    contact_name: str
    contact_id: str
    account_id: str
    updated_on: str
    row_number: int


def load_env_file(path: Path) -> None:
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key or key in os.environ:
            continue

        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]

        os.environ[key] = value


def normalize_whitespace(value: str) -> str:
    return " ".join(value.split())


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKC", value or "")
    value = value.replace("\u00A0", " ")
    value = value.strip().lower()
    value = re.sub(r"[^\w\s]", " ", value)
    value = normalize_whitespace(value)
    return value


def contact_name_candidates(value: str) -> list[str]:
    candidates: list[str] = []
    raw = normalize_text(value)
    if raw:
        candidates.append(raw)

    if "," in value:
        parts = [normalize_whitespace(part) for part in value.split(",") if part.strip()]
        if len(parts) >= 2:
            reordered = normalize_text(" ".join(parts[1:] + [parts[0]]))
            if reordered and reordered not in candidates:
                candidates.append(reordered)

    return candidates


def call_center_candidates(value: str) -> list[str]:
    normalized = normalize_text(value)
    if not normalized:
        return []

    candidates = [normalized]
    alias = SUBAGENT_ALIASES.get(normalized)
    if alias and alias not in candidates:
        candidates.append(alias)

    simplified = re.sub(
        r"\b(bpo|communications|communication|marketing|tech|solutions|solution|comm)\b",
        "",
        normalized,
    )
    simplified = normalize_whitespace(simplified)
    if simplified and simplified not in candidates:
        candidates.append(simplified)

    return candidates


def parse_datetime(value: str) -> tuple[int, str]:
    text = (value or "").strip()
    if not text:
        return (0, "")

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            parsed = datetime.strptime(text, fmt)
            return (int(parsed.timestamp()), text)
        except ValueError:
            continue

    return (0, text)


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def http_json(url: str, headers: dict[str, str]) -> Any:
    request = Request(url, headers=headers)
    with urlopen(request, timeout=60) as response:
        body = response.read().decode("utf-8")
        if not body.strip():
            return None
        return json.loads(body)


def http_json_request(
    url: str,
    headers: dict[str, str],
    method: str,
    payload: Any,
) -> Any:
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = Request(url, headers=headers, data=data, method=method)
    with urlopen(request, timeout=60) as response:
        body = response.read().decode("utf-8")
        if not body.strip():
            return None
        return json.loads(body)


def build_ghl_headers(token: str, version: str, location_id: str) -> dict[str, str]:
    headers = {
        "Authorization": f"Bearer {token}",
        "Version": version,
        "Accept": "application/json",
        "User-Agent": os.environ.get("GHL_API_USER_AGENT", DEFAULT_GHL_USER_AGENT),
        "Accept-Language": "en-US,en;q=0.9",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Origin": "https://app.gohighlevel.com",
        "Referer": "https://app.gohighlevel.com/",
    }
    if location_id:
        headers["Location-Id"] = location_id
    return headers


def fetch_supabase_deals(
    supabase_url: str,
    service_role_key: str,
    page_size: int,
    active_only: bool,
) -> list[DealRecord]:
    print("[ghl-notes] loading deals from Supabase")
    select = "monday_item_id,ghl_name,call_center"
    base = f"{supabase_url.rstrip('/')}/rest/v1/monday_com_deals?select={quote(select)}"
    if active_only:
        base += "&is_active=eq.true"

    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Accept": "application/json",
        "Range-Unit": "items",
    }

    results: list[DealRecord] = []
    offset = 0
    while True:
        range_headers = dict(headers)
        range_headers["Range"] = f"{offset}-{offset + page_size - 1}"
        data = http_json(base, range_headers)
        if not data:
            break

        for row in data:
            ghl_name = str(row.get("ghl_name") or "").strip()
            call_center = str(row.get("call_center") or "").strip()
            monday_item_id = str(row.get("monday_item_id") or "").strip()
            if not monday_item_id or not ghl_name or not call_center:
                continue
            results.append(
                DealRecord(
                    monday_item_id=monday_item_id,
                    ghl_name=ghl_name,
                    call_center=call_center,
                )
            )

        print(f"[ghl-notes] fetched {len(data)} rows from Supabase, total={len(results)}")
        if len(data) < page_size:
            break
        offset += page_size

    return results


def load_deals_from_csv(path: Path) -> list[DealRecord]:
    print(f"[ghl-notes] loading deals from csv {path}")
    results: list[DealRecord] = []
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            monday_item_id = str(
                row.get("monday_item_id")
                or row.get("item_id")
                or ""
            ).strip()
            ghl_name = str(row.get("ghl_name") or row.get("GHL Name") or "").strip()
            call_center = str(row.get("call_center") or row.get("Call Center") or "").strip()
            if not monday_item_id or not ghl_name or not call_center:
                continue
            results.append(
                DealRecord(
                    monday_item_id=monday_item_id,
                    ghl_name=ghl_name,
                    call_center=call_center,
                )
            )

    print(f"[ghl-notes] loaded {len(results)} deals from csv")
    return results


def load_subagents(path: Path) -> tuple[dict[str, SubagentRecord], list[SubagentRecord]]:
    print(f"[ghl-notes] loading subagents from {path}")
    parsed = load_json(path)
    if not isinstance(parsed, list):
        raise RuntimeError("subagent json must contain a top-level array")

    by_name: dict[str, SubagentRecord] = {}
    ordered: list[SubagentRecord] = []
    for entry in parsed:
        if not isinstance(entry, dict):
            continue

        name = str(entry.get("name") or "").strip()
        account_id = str(entry.get("id") or "").strip()
        location_id = str(entry.get("location_id") or "").strip()
        access_token = str(entry.get("access_token") or "").strip()
        api_key = str(entry.get("api_key") or "").strip()
        token = access_token or api_key
        token_source = "access_token" if access_token else "api_key"

        if not name or not account_id or not token:
            continue

        record = SubagentRecord(
            id=account_id,
            name=name,
            location_id=location_id,
            token=token,
            token_source=token_source,
        )
        ordered.append(record)
        by_name[normalize_text(name)] = record

    print(f"[ghl-notes] loaded {len(ordered)} subagents")
    return by_name, ordered


def choose_best_contact(records: list[ContactRecord]) -> ContactRecord:
    def sort_key(record: ContactRecord) -> tuple[int, str, int]:
        timestamp, updated_text = parse_datetime(record.updated_on)
        return (timestamp, updated_text, record.row_number)

    return sorted(records, key=sort_key, reverse=True)[0]


def load_contacts(path: Path) -> dict[tuple[str, str], ContactRecord]:
    print(f"[ghl-notes] loading opportunities workbook {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = sheet.iter_rows(min_row=1, values_only=True)
    headers = [str(cell or "").strip() for cell in next(rows)]
    header_index = {header: index for index, header in enumerate(headers)}
    required = {"Contact Name", "Contact ID", "Account Id", "Updated on"}
    missing = required - set(header_index)
    if missing:
        raise RuntimeError(f"workbook missing required columns: {', '.join(sorted(missing))}")

    grouped: dict[tuple[str, str], list[ContactRecord]] = defaultdict(list)
    row_number = 1
    for values in rows:
        row_number += 1
        contact_name = str(values[header_index["Contact Name"]] or "").strip()
        contact_id = str(values[header_index["Contact ID"]] or "").strip()
        account_id = str(values[header_index["Account Id"]] or "").strip()
        updated_on = str(values[header_index["Updated on"]] or "").strip()
        if not contact_name or not contact_id or not account_id:
            continue

        record = ContactRecord(
            contact_name=contact_name,
            contact_id=contact_id,
            account_id=account_id,
            updated_on=updated_on,
            row_number=row_number,
        )

        for candidate in contact_name_candidates(contact_name):
            grouped[(account_id, candidate)].append(record)

    lookup = {
        key: choose_best_contact(records)
        for key, records in grouped.items()
    }
    print(f"[ghl-notes] indexed {len(lookup)} account/contact pairs from workbook")
    return lookup


def resolve_subagent(
    call_center: str,
    subagents_by_name: dict[str, SubagentRecord],
    ordered_subagents: list[SubagentRecord],
) -> tuple[SubagentRecord | None, str]:
    for candidate in call_center_candidates(call_center):
        if candidate in subagents_by_name:
            return subagents_by_name[candidate], "exact_or_alias"

    normalized_call_center = normalize_text(call_center)
    best: SubagentRecord | None = None
    best_score = -1
    for record in ordered_subagents:
        normalized_subagent = normalize_text(record.name)
        if not normalized_subagent:
            continue
        if normalized_subagent in normalized_call_center or normalized_call_center in normalized_subagent:
            score = min(len(normalized_subagent), len(normalized_call_center))
            if score > best_score:
                best = record
                best_score = score

    if best is not None:
        return best, "contains"

    return None, "not_found"


def fetch_contact_notes(
    contact_id: str,
    token: str,
    location_id: str,
    url_template: str,
    version: str,
) -> Any:
    url = url_template.replace("{contact_id}", quote(contact_id))
    return http_json(url, build_ghl_headers(token, version, location_id))


def extract_note_list(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [entry for entry in payload if isinstance(entry, dict)]
    if isinstance(payload, dict):
        for key in ("notes", "data", "contactNotes"):
            value = payload.get(key)
            notes = extract_note_list(value)
            if notes:
                return notes
    return []


def extract_note_count(payload: Any) -> int | None:
    if isinstance(payload, list):
        return len(payload)
    if isinstance(payload, dict):
        for key in ("notes", "data", "contactNotes"):
            value = payload.get(key)
            count = extract_note_count(value)
            if count is not None:
                return count
    return None


def choose_note(notes: list[dict[str, Any]]) -> dict[str, Any] | None:
    def sort_key(entry: dict[str, Any]) -> tuple[int, str]:
        for key in ("updatedAt", "updated_at", "createdAt", "created_at", "dateAdded"):
            value = str(entry.get(key) or "").strip()
            timestamp, text = parse_datetime(value.replace("T", " ").replace("Z", ""))
            if timestamp or text:
                return (timestamp, text)
        return (0, "")

    if not notes:
        return None
    return sorted(notes, key=sort_key, reverse=True)[0]


def current_timestamp_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def build_supabase_notes_row(result: dict[str, Any], fetched_at: str) -> dict[str, Any]:
    return {
        "monday_item_id": result["monday_item_id"],
        "ghl_name": result["ghl_name"],
        "call_center": result["call_center"],
        "status": result["status"],
        "subagent_match_mode": result["subagent_match_mode"],
        "subagent_name": result["subagent_name"],
        "subagent_account_id": result["account_id"],
        "contact_name": result["contact_name"],
        "contact_id": result["contact_id"],
        "notes_count": result["notes_count"],
        "latest_note_id": result["latest_note_id"],
        "latest_note_summary": result["latest_note_summary"],
        "notes": result["notes"],
        "notes_payload": result["notes_payload"],
        "notes_error": result["notes_error"],
        "fetched_at": fetched_at,
        "updated_at": fetched_at,
    }


def upsert_results_to_supabase(
    supabase_url: str,
    service_role_key: str,
    table_name: str,
    rows: list[dict[str, Any]],
    batch_size: int,
) -> None:
    if not rows:
        print("[ghl-notes] no rows to upsert into Supabase")
        return

    endpoint = (
        f"{supabase_url.rstrip('/')}/rest/v1/{quote(table_name)}"
        f"?on_conflict={quote('monday_item_id')}"
    )
    headers = {
        "apikey": service_role_key,
        "Authorization": f"Bearer {service_role_key}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }

    total = len(rows)
    safe_batch_size = max(batch_size, 1)
    for start in range(0, total, safe_batch_size):
        batch = rows[start:start + safe_batch_size]
        batch_number = (start // safe_batch_size) + 1
        print(
            f"[ghl-notes] upserting supabase batch {batch_number} "
            f"size={len(batch)} table={table_name}"
        )
        http_json_request(
            endpoint,
            headers=headers,
            method="POST",
            payload=batch,
        )
        print(f"[ghl-notes] upserted {min(start + len(batch), total)}/{total} rows")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Fetch all GHL contact note lists for monday_com_deals")
    parser.add_argument("--subagents-json", default="notes/subagent.json")
    parser.add_argument("--contacts-xlsx", default="notes/GHL_Opportunities_New_API_Export (19).xlsx")
    parser.add_argument("--output-json", default="notes/deal_contact_notes.json")
    parser.add_argument("--output-csv", default="notes/deal_contact_notes.csv")
    parser.add_argument("--deals-csv", help="Optional CSV fallback for deals instead of querying Supabase")
    parser.add_argument("--page-size", type=int, default=1000)
    parser.add_argument("--active-only", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit", type=int, default=0, help="Optional limit on matched deals to process")
    parser.add_argument("--sleep-ms", type=int, default=100)
    parser.add_argument(
        "--notes-url-template",
        default=os.environ.get(
            "GHL_CONTACT_NOTES_URL_TEMPLATE",
            "https://services.leadconnectorhq.com/contacts/{contact_id}/notes",
        ),
    )
    parser.add_argument(
        "--ghl-version",
        default=os.environ.get("GHL_API_VERSION", "2021-07-28"),
    )
    parser.add_argument(
        "--supabase-table",
        default=os.environ.get("GHL_NOTES_SUPABASE_TABLE", "monday_deal_contact_notes"),
    )
    parser.add_argument("--supabase-upsert-batch-size", type=int, default=200)
    parser.add_argument("--skip-supabase-upsert", action="store_true")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    load_env_file((Path.cwd() / ".env").resolve())

    root = Path.cwd()
    subagents_path = (root / args.subagents_json).resolve()
    contacts_path = (root / args.contacts_xlsx).resolve()
    output_json = (root / args.output_json).resolve()
    output_csv = (root / args.output_csv).resolve()

    subagents_by_name, ordered_subagents = load_subagents(subagents_path)
    contacts_lookup = load_contacts(contacts_path)

    if args.deals_csv:
        deals = load_deals_from_csv((root / args.deals_csv).resolve())
    else:
        supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("SUPABASE_URL")
        service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        if not supabase_url or not service_role_key:
            raise RuntimeError(
                "NEXT_PUBLIC_SUPABASE_URL/SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required unless --deals-csv is provided"
            )
        deals = fetch_supabase_deals(
            supabase_url=supabase_url,
            service_role_key=service_role_key,
            page_size=args.page_size,
            active_only=args.active_only,
        )

    unique_deals: list[DealRecord] = []
    seen_keys: set[tuple[str, str]] = set()
    for deal in deals:
        key = (normalize_text(deal.ghl_name), normalize_text(deal.call_center))
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique_deals.append(deal)

    print(
        f"[ghl-notes] deals loaded={len(deals)} unique ghl_name/call_center pairs={len(unique_deals)}"
    )

    results: list[dict[str, Any]] = []
    notes_cache: dict[tuple[str, str], dict[str, Any]] = {}
    counters: Counter[str] = Counter()

    for index, deal in enumerate(unique_deals, start=1):
        if args.limit and len(results) >= args.limit:
            break

        subagent, subagent_match_mode = resolve_subagent(
            deal.call_center,
            subagents_by_name,
            ordered_subagents,
        )

        result: dict[str, Any] = {
            "monday_item_id": deal.monday_item_id,
            "ghl_name": deal.ghl_name,
            "call_center": deal.call_center,
            "status": "",
            "subagent_match_mode": subagent_match_mode,
            "subagent_name": None,
            "account_id": None,
            "location_id": None,
            "contact_name": None,
            "contact_id": None,
            "notes_count": None,
            "latest_note_id": None,
            "latest_note_summary": None,
            "notes_error": None,
            "notes_payload": None,
            "notes": None,
        }

        if subagent is None:
            result["status"] = "subagent_not_found"
            counters[result["status"]] += 1
            results.append(result)
            continue

        result["subagent_name"] = subagent.name
        result["account_id"] = subagent.id
        result["location_id"] = subagent.location_id

        matched_contact: ContactRecord | None = None
        for candidate in contact_name_candidates(deal.ghl_name):
            matched_contact = contacts_lookup.get((subagent.id, candidate))
            if matched_contact is not None:
                break

        if matched_contact is None:
            result["status"] = "contact_not_found"
            counters[result["status"]] += 1
            results.append(result)
            continue

        result["contact_name"] = matched_contact.contact_name
        result["contact_id"] = matched_contact.contact_id

        cache_key = (matched_contact.contact_id, subagent.token)
        if args.dry_run:
            result["status"] = "dry_run"
            counters[result["status"]] += 1
            results.append(result)
            continue

        if cache_key not in notes_cache:
            try:
                payload = fetch_contact_notes(
                    contact_id=matched_contact.contact_id,
                    token=subagent.token,
                    location_id=subagent.location_id,
                    url_template=args.notes_url_template,
                    version=args.ghl_version,
                )
                notes_cache[cache_key] = {
                    "payload": payload,
                    "notes": extract_note_list(payload),
                    "error": None,
                    "notes_count": extract_note_count(payload),
                }
                time.sleep(max(args.sleep_ms, 0) / 1000)
            except HTTPError as exc:
                body = exc.read().decode("utf-8", errors="replace")
                notes_cache[cache_key] = {
                    "payload": None,
                    "notes": None,
                    "error": f"HTTP {exc.code}: {body}",
                    "notes_count": None,
                }
            except URLError as exc:
                notes_cache[cache_key] = {
                    "payload": None,
                    "notes": None,
                    "error": f"URL error: {exc}",
                    "notes_count": None,
                }
            except Exception as exc:  # pragma: no cover - defensive runtime error handling
                notes_cache[cache_key] = {
                    "payload": None,
                    "notes": None,
                    "error": str(exc),
                    "notes_count": None,
                }

        cache_entry = notes_cache[cache_key]
        latest_note = choose_note(cache_entry["notes"] or [])
        result["notes_payload"] = cache_entry["payload"]
        result["notes"] = cache_entry["notes"]
        result["notes_error"] = cache_entry["error"]
        result["notes_count"] = cache_entry["notes_count"]
        result["latest_note_summary"] = latest_note
        result["latest_note_id"] = str(latest_note.get("id") or "") if latest_note else None
        result["status"] = "notes_fetched" if cache_entry["error"] is None else "notes_fetch_failed"
        counters[result["status"]] += 1
        results.append(result)

        if index % 100 == 0:
            print(f"[ghl-notes] processed {index}/{len(unique_deals)} unique deals")

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_csv.parent.mkdir(parents=True, exist_ok=True)

    output_json.write_text(json.dumps(results, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    csv_fields = [
        "monday_item_id",
        "ghl_name",
        "call_center",
        "status",
        "subagent_match_mode",
        "subagent_name",
        "account_id",
        "location_id",
        "contact_name",
        "contact_id",
        "notes_count",
        "latest_note_id",
        "notes_error",
    ]
    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=csv_fields)
        writer.writeheader()
        for row in results:
            writer.writerow({field: row.get(field) for field in csv_fields})

    print(f"[ghl-notes] wrote json -> {output_json}")
    print(f"[ghl-notes] wrote csv -> {output_csv}")

    if args.dry_run:
        print("[ghl-notes] skipping Supabase upsert because --dry-run was used")
    elif args.skip_supabase_upsert:
        print("[ghl-notes] skipping Supabase upsert because --skip-supabase-upsert was used")
    else:
        supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or os.environ.get("SUPABASE_URL")
        service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        if not supabase_url or not service_role_key:
            raise RuntimeError(
                "NEXT_PUBLIC_SUPABASE_URL/SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required for Supabase upsert"
            )

        fetched_at = current_timestamp_iso()
        upsert_rows = [build_supabase_notes_row(result, fetched_at) for result in results]
        upsert_results_to_supabase(
            supabase_url=supabase_url,
            service_role_key=service_role_key,
            table_name=args.supabase_table,
            rows=upsert_rows,
            batch_size=args.supabase_upsert_batch_size,
        )

    print("[ghl-notes] summary")
    for status, count in sorted(counters.items()):
        print(f"  - {status}: {count}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
